import io
import csv
from datetime import datetime, date, timedelta
from decimal import Decimal

from django.conf import settings
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Sum, Count, Q, F, Avg, FloatField, ExpressionWrapper
from django.db.models.functions import TruncMonth, ExtractMonth
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.translation import gettext_lazy as _

from apps.projects.models import Project
from apps.invoices.models import Invoice, Payment
from apps.inventory.models import Material, StockTransaction
from apps.employees.models import Employee, Attendance
from apps.clients.models import Client

from .forms import ReportFilterForm


def is_accountant_or_admin(user):
    return user.role in ('admin', 'accountant')


@login_required
def report_dashboard(request):
    today = timezone.now().date()
    first_of_month = today.replace(day=1)

    project_count = Project.objects.count()
    active_projects = Project.objects.filter(status='Ongoing').count()
    completed_projects = Project.objects.filter(status='Completed').count()

    total_invoiced = Invoice.objects.aggregate(total=Sum('grand_total'))['total'] or 0
    total_paid = Payment.objects.aggregate(total=Sum('amount'))['total'] or 0
    outstanding = total_invoiced - total_paid

    low_stock = Material.objects.filter(quantity__lte=F('reorder_level')).count()
    total_employees = Employee.objects.count()
    pending_attendance = Attendance.objects.filter(date=today, status='absent').count()

    context = {
        'title': _('Reports Dashboard'),
        'project_count': project_count,
        'active_projects': active_projects,
        'completed_projects': completed_projects,
        'total_invoiced': total_invoiced,
        'total_paid': total_paid,
        'outstanding': outstanding,
        'low_stock': low_stock,
        'total_employees': total_employees,
        'pending_attendance': pending_attendance,
    }
    return render(request, 'reports/report_dashboard.html', context)


@login_required
def financial_report(request):
    today = timezone.now().date()
    first_of_month = today.replace(day=1)

    date_from = request.GET.get('date_from', first_of_month.isoformat())
    date_to = request.GET.get('date_to', today.isoformat())
    project_id = request.GET.get('project', '')

    filters = Q()
    if date_from:
        filters &= Q(date__gte=date_from)
    if date_to:
        filters &= Q(date__lte=date_to)
    if project_id:
        filters &= Q(project_id=project_id)

    invoices = Invoice.objects.filter(filters).select_related('project', 'client')
    payments = Payment.objects.filter(
        Q(invoice__date__gte=date_from) if date_from else Q(),
        Q(invoice__date__lte=date_to) if date_to else Q(),
    ).select_related('invoice', 'invoice__project')

    if project_id:
        payments = payments.filter(invoice__project_id=project_id)

    total_revenue = invoices.aggregate(total=Sum('grand_total'))['total'] or 0
    total_collected = payments.aggregate(total=Sum('amount'))['total'] or 0
    total_outstanding = total_revenue - total_collected

    invoices_paid = invoices.filter(status='paid').count()
    invoices_pending = invoices.filter(status='pending').count()
    invoices_overdue = invoices.filter(status='overdue').count()

    monthly_data = []
    months_data = invoices.annotate(
        month=TruncMonth('date')
    ).values('month').annotate(
        total=Sum('grand_total')
    ).order_by('month')

    for entry in months_data:
        if entry['month']:
            monthly_data.append({
                'month': entry['month'].strftime('%b %Y'),
                'total': float(entry['total'] or 0),
            })

    project_choices = list(Project.objects.values_list('id', 'name'))
    form = ReportFilterForm(
        initial={'date_from': date_from, 'date_to': date_to, 'project': project_id},
        project_choices=project_choices,
    )

    context = {
        'title': _('Financial Report'),
        'invoices': invoices,
        'total_revenue': total_revenue,
        'total_collected': total_collected,
        'total_outstanding': total_outstanding,
        'invoices_paid': invoices_paid,
        'invoices_pending': invoices_pending,
        'invoices_overdue': invoices_overdue,
        'monthly_data': monthly_data,
        'form': form,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'reports/financial_report.html', context)


@login_required
def project_report(request):
    status_filter = request.GET.get('status', '')
    project_id = request.GET.get('project', '')

    projects = Project.objects.all().select_related('client')

    if status_filter:
        projects = projects.filter(status=status_filter)
    if project_id:
        projects = projects.filter(id=project_id)

    total_projects = projects.count()
    completed = projects.filter(status='Completed').count()
    active = projects.filter(status='Ongoing').count()
    on_hold = projects.filter(status='On Hold').count()
    cancelled = projects.filter(status='Cancelled').count()

    total_budget = projects.aggregate(total=Sum('budget'))['total'] or 0
    total_spent = 0

    for project in projects:
        project.completion_pct = project.progress_percent or 0
        if project.budget and project.budget > 0:
            project.budget_used_pct = 0
        else:
            project.budget_used_pct = 0

    project_choices = list(Project.objects.values_list('id', 'name'))
    form = ReportFilterForm(
        initial={'project': project_id},
        project_choices=project_choices,
    )

    context = {
        'title': _('Project Report'),
        'projects': projects,
        'total_projects': total_projects,
        'completed': completed,
        'active': active,
        'on_hold': on_hold,
        'cancelled': cancelled,
        'total_budget': total_budget,
        'total_spent': total_spent,
        'form': form,
        'status_filter': status_filter,
    }
    return render(request, 'reports/project_report.html', context)


@login_required
def inventory_report(request):
    category = request.GET.get('category', '')
    stock_status = request.GET.get('stock_status', '')

    materials = Material.objects.all()

    if category:
        materials = materials.filter(category=category)
    if stock_status == 'low':
        materials = materials.filter(quantity__lte=F('reorder_level'))
    elif stock_status == 'out':
        materials = materials.filter(quantity=0)
    elif stock_status == 'in':
        materials = materials.filter(quantity__gt=0)

    total_materials = materials.count()
    low_stock_count = materials.filter(quantity__lte=F('reorder_level')).count()
    out_of_stock = materials.filter(quantity=0).count()
    total_stock_value = materials.aggregate(
        total=Sum(F('quantity') * F('unit_cost'), output_field=FloatField())
    )['total'] or 0

    recent_transactions = StockTransaction.objects.all().select_related(
        'material', 'done_by'
    ).order_by('-created_at')[:50]

    for material in materials:
        material.stock_value = (material.quantity or 0) * (material.unit_cost or 0)
        if material.reorder_level and material.reorder_level > 0:
            material.stock_status = 'low' if material.quantity <= material.reorder_level else 'adequate'
        else:
            material.stock_status = 'adequate'
        if material.quantity == 0:
            material.stock_status = 'out'

    context = {
        'title': _('Inventory Report'),
        'materials': materials,
        'total_materials': total_materials,
        'low_stock_count': low_stock_count,
        'out_of_stock': out_of_stock,
        'total_stock_value': total_stock_value,
        'recent_transactions': recent_transactions,
        'category': category,
        'stock_status': stock_status,
    }
    return render(request, 'reports/inventory_report.html', context)


@login_required
def employee_report(request):
    today = timezone.now().date()
    first_of_month = today.replace(day=1)

    date_from = request.GET.get('date_from', first_of_month.isoformat())
    date_to = request.GET.get('date_to', today.isoformat())

    employees = Employee.objects.all().select_related('user')

    date_filter = Q(date__gte=date_from, date__lte=date_to)

    for emp in employees:
        attendance_records = Attendance.objects.filter(
            employee=emp
        ).filter(date_filter)

        total_days = attendance_records.count()
        present = attendance_records.filter(status='present').count()
        absent = attendance_records.filter(status='absent').count()
        late = attendance_records.filter(status='late').count()
        half_day = attendance_records.filter(status='half_day').count()

        emp.total_days = total_days
        emp.present_days = present
        emp.absent_days = absent
        emp.late_days = late
        emp.half_days = half_day
        emp.attendance_rate = round((present / total_days * 100) if total_days > 0 else 0, 1)

    total_employees = employees.count()
    overall_present = Attendance.objects.filter(date_filter, status='present').count()
    overall_absent = Attendance.objects.filter(date_filter, status='absent').count()
    overall_late = Attendance.objects.filter(date_filter, status='late').count()

    context = {
        'title': _('Employee Report'),
        'employees': employees,
        'total_employees': total_employees,
        'overall_present': overall_present,
        'overall_absent': overall_absent,
        'overall_late': overall_late,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'reports/employee_report.html', context)


@login_required
def export_pdf(request, report_type):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, mm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, Image, HRFlowable,
        )
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    except ImportError:
        return HttpResponse(
            _('ReportLab is not installed. Please install it to generate PDFs.'),
            content_type='text/plain',
        )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='Title2',
        parent=styles['Title'],
        fontSize=18,
        textColor=colors.HexColor('#1a3c6e'),
        spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        name='SubTitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#666666'),
        spaceAfter=20,
    ))
    styles.add(ParagraphStyle(
        name='TableHeader',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.white,
        backColor=colors.HexColor('#1a3c6e'),
    ))

    elements = []
    company_name = getattr(settings, 'COMPANY_NAME', 'SCHONES HEIM BUILDERS')
    elements.append(Paragraph(company_name, styles['Title2']))
    elements.append(Paragraph(
        f'{report_type.replace("_", " ").title()} Report - Generated {datetime.now().strftime("%d %b %Y %H:%M")}',
        styles['SubTitle'],
    ))
    elements.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#1a3c6e')))
    elements.append(Spacer(1, 10))

    if report_type == 'financial':
        invoices = Invoice.objects.all().select_related('project')[:100]
        data = [[_('Invoice #'), _('Project'), _('Date'), _('Amount'), _('Status')]]
        for inv in invoices:
            data.append([
                inv.invoice_number or str(inv.id),
                inv.project.name if inv.project else '-',
                inv.date.strftime('%d/%m/%Y') if inv.date else '-',
                f'{inv.grand_total:,.2f}',
                inv.get_status_display(),
            ])
        if len(data) > 1:
            data.append(['', '', _('Total'), f'{sum(i.grand_total for i in invoices):,.2f}', ''])

        table = Table(data, colWidths=[80, 120, 80, 80, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3c6e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f4f8')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(table)

    elif report_type == 'project':
        projects = Project.objects.all().select_related('client')[:100]
        data = [[_('Project'), _('Client'), _('Budget'), _('Actual'), _('Progress'), _('Status')]]
        for p in projects:
            data.append([
                p.name,
                str(p.client) if p.client else '-',
                f'{p.budget:,.2f}' if p.budget else '-',
                '-',
                f'{p.progress_percent or 0}%',
                p.get_status_display() if p.status else '-',
            ])
        table = Table(data, colWidths=[120, 100, 80, 80, 70, 70])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3c6e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (2, 0), (4, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table)

    elif report_type == 'inventory':
        materials = Material.objects.all()[:100]
        data = [[_('Material'), _('Category'), _('Qty'), _('Unit Price'), _('Value')]]
        for m in materials:
            val = (m.quantity or 0) * (m.unit_cost or 0)
            data.append([
                m.name,
                m.get_category_display() if m.category else '-',
                str(m.quantity or 0),
                f'{m.unit_cost:,.2f}' if m.unit_cost else '-',
                f'{val:,.2f}',
            ])
        total_val = sum((m.quantity or 0) * (m.unit_cost or 0) for m in materials)
        data.append(['', '', '', _('Total Value'), f'{total_val:,.2f}'])
        table = Table(data, colWidths=[110, 90, 70, 80, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3c6e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (2, 0), (4, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f4f8')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(table)

    elif report_type == 'employee':
        employees = Employee.objects.all().select_related('user')[:100]
        data = [[_('Name'), _('Position'), _('Department'), _('Phone'), _('Status')]]
        for emp in employees:
            data.append([
                emp.user.get_full_name() if emp.user else '-',
                emp.position or '-',
                emp.department or '-',
                emp.user.phone if emp.user else '-',
                'Active' if emp.is_active else 'Inactive',
            ])
        table = Table(data, colWidths=[120, 90, 90, 90, 60])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3c6e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table)

    elements.append(Spacer(1, 20))
    elements.append(HRFlowable(width='100%', thickness=0.5, color=colors.grey))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(
        f'&copy; {datetime.now().year} {company_name}. All rights reserved.',
        ParagraphStyle('Footer', fontSize=8, textColor=colors.grey, alignment=TA_CENTER),
    ))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="{report_type}_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
    )
    return response


@login_required
def export_excel(request, report_type):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            _('openpyxl is not installed. Please install it to generate Excel files.'),
            content_type='text/plain',
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type.replace('_', ' ').title()

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A3C6E', end_color='1A3C6E', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    company_row = ws.cell(row=1, column=1, value='SCHONES HEIM BUILDERS')
    company_row.font = Font(name='Calibri', bold=True, size=14, color='1A3C6E')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    title_row = ws.cell(
        row=2, column=1,
        value=f'{report_type.replace("_", " ").title()} Report - {datetime.now().strftime("%d %b %Y")}',
    )
    title_row.font = Font(name='Calibri', size=10, color='666666')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    start_row = 4

    if report_type == 'financial':
        headers = [str(_('Invoice #')), str(_('Project')), str(_('Date')), str(_('Amount')), str(_('Status'))]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        invoices = Invoice.objects.all().select_related('project')[:500]
        for i, inv in enumerate(invoices, start_row + 1):
            ws.cell(row=i, column=1, value=inv.invoice_number or str(inv.id)).border = thin_border
            ws.cell(row=i, column=2, value=inv.project.name if inv.project else '').border = thin_border
            ws.cell(row=i, column=3, value=inv.date.strftime('%d/%m/%Y') if inv.date else '').border = thin_border
            c = ws.cell(row=i, column=4, value=float(inv.grand_total))
            c.number_format = '#,##0.00'
            c.border = thin_border
            ws.cell(row=i, column=5, value=inv.get_status_display()).border = thin_border

        total_row = start_row + 1 + invoices.count()
        ws.cell(row=total_row, column=3, value=str(_('Total'))).font = Font(bold=True)
        c = ws.cell(row=total_row, column=4, value=float(sum(i.grand_total for i in invoices)))
        c.font = Font(bold=True)
        c.number_format = '#,##0.00'
        for col in range(1, 6):
            ws.cell(row=total_row, column=col).border = thin_border
            ws.cell(row=total_row, column=col).fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

    elif report_type == 'project':
        headers = [str(_('Project')), str(_('Client')), str(_('Budget')), str(_('Actual')), str(_('Progress')), str(_('Status'))]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        projects = Project.objects.all().select_related('client')[:500]
        for i, p in enumerate(projects, start_row + 1):
            ws.cell(row=i, column=1, value=p.name).border = thin_border
            ws.cell(row=i, column=2, value=str(p.client) if p.client else '').border = thin_border
            c = ws.cell(row=i, column=3, value=float(p.budget) if p.budget else 0)
            c.number_format = '#,##0.00'
            c.border = thin_border
            c = ws.cell(row=i, column=4, value=0)
            c.number_format = '#,##0.00'
            c.border = thin_border
            ws.cell(row=i, column=5, value=f'{p.progress_percent or 0}%').border = thin_border
            ws.cell(row=i, column=6, value=p.get_status_display()).border = thin_border

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15

    elif report_type == 'inventory':
        headers = [str(_('Material')), str(_('Category')), str(_('Quantity')), str(_('Unit Price')), str(_('Stock Value'))]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        materials = Material.objects.all()[:500]
        for i, m in enumerate(materials, start_row + 1):
            ws.cell(row=i, column=1, value=m.name).border = thin_border
            ws.cell(row=i, column=2, value=m.get_category_display() if m.category else '').border = thin_border
            c = ws.cell(row=i, column=3, value=float(m.quantity or 0))
            c.number_format = '#,##0'
            c.border = thin_border
            c = ws.cell(row=i, column=4, value=float(m.unit_cost or 0))
            c.number_format = '#,##0.00'
            c.border = thin_border
            val = (m.quantity or 0) * (m.unit_cost or 0)
            c = ws.cell(row=i, column=5, value=float(val))
            c.number_format = '#,##0.00'
            c.border = thin_border

        total_row = start_row + 1 + materials.count()
        ws.cell(row=total_row, column=4, value=str(_('Total Value'))).font = Font(bold=True)
        total_val = sum((m.quantity or 0) * (m.unit_cost or 0) for m in materials)
        c = ws.cell(row=total_row, column=5, value=float(total_val))
        c.font = Font(bold=True)
        c.number_format = '#,##0.00'
        for col in range(1, 6):
            ws.cell(row=total_row, column=col).border = thin_border
            ws.cell(row=total_row, column=col).fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 16

    elif report_type == 'employee':
        headers = [str(_('Name')), str(_('Position')), str(_('Department')), str(_('Phone')), str(_('Status'))]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        employees = Employee.objects.all().select_related('user')[:500]
        for i, emp in enumerate(employees, start_row + 1):
            ws.cell(row=i, column=1, value=emp.user.get_full_name() if emp.user else '').border = thin_border
            ws.cell(row=i, column=2, value=emp.position or '').border = thin_border
            ws.cell(row=i, column=3, value=emp.department or '').border = thin_border
            ws.cell(row=i, column=4, value=emp.user.phone if emp.user else '').border = thin_border
            ws.cell(row=i, column=5, value='Active' if emp.is_active else 'Inactive').border = thin_border

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12

    ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width or 8, 12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{report_type}_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    )
    wb.save(response)
    return response
